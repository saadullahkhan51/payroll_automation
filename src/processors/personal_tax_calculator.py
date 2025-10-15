import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
from typing import List, Dict
from decimal import Decimal
from datetime import date
from database.repository import PayrollRepository
from database.models import PayrollRecordDB
from config.settings import OUTPUT_DIR
import calendar
import os

class PersonalTaxCalculator:
    """Process and generate annual summaries"""
    
    def __init__(self, repository: PayrollRepository):
        self.repo = repository
        self.output_dir = OUTPUT_DIR / "personal_tax"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_personal_annual_summary(self, employee_id: str, year: int) -> str:
        """Generate personal annual tax calculations Excel matching the format"""
        
        # Get payroll records
        records = self.repo.get_payroll_records(employee_id, year)
        
        if not records:
            raise ValueError(f"No payroll records found for employee {employee_id} in {year}")
        
        # Get employee info
        employee = self.repo.get_employee(employee_id)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year} Tax Calc"
        
        # Set column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 18
        
        # Define styles
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        
        # Title row
        row = 1
        ws[f'A{row}'] = f"{employee.name}, worker"
        ws[f'A{row}'].font = bold_font
        
        # Headers row
        row = 2
        ws[f'A{row}'] = "PALKKAJAKSO"
        ws[f'B{row}'] = "Ruotsi (laskelmanro)"
        ws[f'C{row}'] = "Taxable income SEK"
        ws[f'D{row}'] = "Taxable income EUR"
        ws[f'E{row}'] = "Verot Ruotsisin EUR"
        ws[f'F{row}'] = "allowances SEK"
        ws[f'G{row}'] = "allowances EUR"
        ws[f'H{row}'] = "Verot Ruotsisin allowances EUR"
        ws[f'I{row}'] = "Norja (laskelmanro)"
        ws[f'J{row}'] = "NOK"
        ws[f'K{row}'] = "EUR (Norja)"
        ws[f'L{row}'] = "Verot Norjaan EUR"
        
        # Apply header styling
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Exchange rate (approximate SEK to EUR)
        SEK_TO_EUR = Decimal('0.086')  # 1 SEK ≈ 0.086 EUR
        
        # Data rows
        row = 3
        totals = {
            'taxable_income_sek': Decimal('0'),
            'taxable_income_eur': Decimal('0'),
            'tax_eur': Decimal('0'),
            'allowances_sek': Decimal('0'),
            'allowances_eur': Decimal('0'),
            'tax_allowances_eur': Decimal('0')
        }

        month_range = lambda m, y: f"1.-{calendar.monthrange(y, m)[1]}."

        for record in sorted(records, key=lambda r: (r.pay_period_start)):
            # Format period
            period = f"{record.pay_period_start.strftime('%d.%m.%Y')}"
            period_end = f"{record.pay_period_end.strftime('%d.%m.%Y')}"

            # Calculate Swedish calculation number (arbitrary format like "93800")
            calc_num = 93800 + (record.month * 100)
            
            # Calculate taxable income (gross salary converted to SEK)
            taxable_income_eur = record.gross_salary
            taxable_income_sek = taxable_income_eur / SEK_TO_EUR
            
            # Swedish tax in EUR
            swedish_tax_eur = Decimal(str(record.swedish_tax))
            
            # Allowances (tax-free portions)
            allowances_eur = record.tax_free_portion
            allowances_sek = allowances_eur / SEK_TO_EUR
            
            # Tax on allowances (portion of Swedish tax allocated to allowances)
            # Typically 30% of allowances
            tax_allowances_eur = allowances_eur * Decimal('0.30')
            
            # Write data
            ws[f'A{row}'] = period
            ws[f'B{row}'] = calc_num
            ws[f'C{row}'] = float(taxable_income_sek)
            ws[f'D{row}'] = float(taxable_income_eur)
            ws[f'E{row}'] = float(swedish_tax_eur)
            ws[f'F{row}'] = float(allowances_sek)
            ws[f'G{row}'] = float(allowances_eur)
            ws[f'H{row}'] = float(tax_allowances_eur)
            # Columns I-L (Norway) left empty as we're only dealing with Sweden
            
            # Apply borders
            for col in range(1, 13):
                ws.cell(row=row, column=col).border = thin_border
            
            # Number formatting
            for col in ['C', 'F']:  # SEK columns
                ws[f'{col}{row}'].number_format = '#,##0.00'
            for col in ['D', 'E', 'G', 'H']:  # EUR columns
                ws[f'{col}{row}'].number_format = '#,##0.00'
            
            # Update totals
            totals['taxable_income_sek'] += taxable_income_sek
            totals['taxable_income_eur'] += taxable_income_eur
            totals['tax_eur'] += swedish_tax_eur
            totals['allowances_sek'] += allowances_sek
            totals['allowances_eur'] += allowances_eur
            totals['tax_allowances_eur'] += tax_allowances_eur
            
            row += 1
        
        # TOTAL row
        ws[f'A{row}'] = "TOTAL"
        ws[f'A{row}'].font = bold_font
        ws[f'C{row}'] = float(totals['taxable_income_sek'])
        ws[f'D{row}'] = float(totals['taxable_income_eur'])
        ws[f'E{row}'] = float(totals['tax_eur'])
        ws[f'F{row}'] = float(totals['allowances_sek'])
        ws[f'G{row}'] = float(totals['allowances_eur'])
        ws[f'H{row}'] = float(totals['tax_allowances_eur'])
        
        # Apply total row styling
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.font = bold_font
            cell.border = thin_border
        
        for col in ['C', 'F']:
            ws[f'{col}{row}'].number_format = '#,##0.00'
        for col in ['D', 'E', 'G', 'H']:
            ws[f'{col}{row}'].number_format = '#,##0.00'
        
        # Summary section
        row += 1
        ws[f'A{row}'] = "Rahapalkasta maksetut verot"
        ws[f'A{row}'].font = bold_font
        ws[f'E{row}'] = float(totals['tax_eur'])
        ws[f'E{row}'].number_format = '#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "Verottomista korvauksista maksetut verot"
        ws[f'A{row}'].font = bold_font
        ws[f'E{row}'] = float(totals['tax_allowances_eur'])
        ws[f'E{row}'].number_format = '#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "YHTEENSÄ"
        ws[f'A{row}'].font = bold_font
        ws[f'E{row}'] = float(totals['tax_eur'] + totals['tax_allowances_eur'])
        ws[f'E{row}'].number_format = '#,##0.00'
        ws[f'E{row}'].font = bold_font
        
        # Generate filename
        filename = f"{employee_id}_{year}_personal_tax_calculations.xlsx"
        filepath = self.output_dir / filename
        
        # Save workbook
        if not os.path.exists(filepath):
            wb.save(filepath)
        
        return str(filepath)
    
    def generate_all_workers_annual_summary(self, year: int) -> str:
        """Generate annual summary for all workers"""
        
        # Get all employees
        employees = self.repo.get_all_employees()
        
        if not employees:
            raise ValueError(f"No employees found")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year} All Workers"
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        
        # Define styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        total_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        # Title
        row = 1
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f"ANNUAL SUMMARY - ALL WORKERS"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        row = 2
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f"Year: {year}"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        # Headers
        row = 4
        headers = [
            'Employee ID',
            'Name',
            'Total Gross',
            'Total Hours',
            'Swedish Tax',
            'Finnish Tax',
            'Pension',
            'Tax-Free',
            'Net Payment',
            'Avg Monthly'
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        row = 5
        company_totals = {
            'gross': Decimal('0'),
            'hours': Decimal('0'),
            'swedish_tax': Decimal('0'),
            'finnish_tax': Decimal('0'),
            'pension': Decimal('0'),
            'tax_free': Decimal('0'),
            'net': Decimal('0')
        }
        
        for employee in employees:
            records = self.repo.get_payroll_records(employee.id, year)
            
            if not records:
                continue
            
            annual_data = self._calculate_annual_totals(records)
            
            ws[f'A{row}'] = employee.id
            ws[f'B{row}'] = employee.name
            ws[f'C{row}'] = float(annual_data['total_gross'])
            ws[f'D{row}'] = float(annual_data['total_hours'])
            ws[f'E{row}'] = float(annual_data['swedish_tax'])
            ws[f'F{row}'] = float(annual_data['finnish_tax'])
            ws[f'G{row}'] = float(annual_data['pension'])
            ws[f'H{row}'] = float(annual_data['tax_free'])
            ws[f'I{row}'] = float(annual_data['total_net'])
            ws[f'J{row}'] = float(annual_data['total_gross'] / len(records)) if records else 0
            
            # Borders
            for col_idx in range(1, 11):
                ws.cell(row=row, column=col_idx).border = thin_border
            
            # Number formatting
            for col in ['C', 'E', 'F', 'G', 'H', 'I', 'J']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
            ws[f'D{row}'].number_format = '#,##0.00'
            
            # Update company totals
            company_totals['gross'] += annual_data['total_gross']
            company_totals['hours'] += annual_data['total_hours']
            company_totals['swedish_tax'] += annual_data['swedish_tax']
            company_totals['finnish_tax'] += annual_data['finnish_tax']
            company_totals['pension'] += annual_data['pension']
            company_totals['tax_free'] += annual_data['tax_free']
            company_totals['net'] += annual_data['total_net']
            
            row += 1
        
        # Company totals
        ws[f'A{row}'] = "COMPANY TOTAL"
        ws[f'A{row}'].font = bold_font
        ws[f'C{row}'] = float(company_totals['gross'])
        ws[f'D{row}'] = float(company_totals['hours'])
        ws[f'E{row}'] = float(company_totals['swedish_tax'])
        ws[f'F{row}'] = float(company_totals['finnish_tax'])
        ws[f'G{row}'] = float(company_totals['pension'])
        ws[f'H{row}'] = float(company_totals['tax_free'])
        ws[f'I{row}'] = float(company_totals['net'])
        
        for col_idx in range(1, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = thin_border
            cell.fill = total_fill
            cell.font = bold_font
        
        for col in ['C', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws[f'{col}{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].number_format = '#,##0.00'
        
        # Generate filename
        filename = f"annual_summary_all_workers_{year}.xlsx"
        filepath = self.output_dir / filename
        
        # Save workbook
        wb.save(filepath)
        
        return str(filepath)
    
    def _calculate_annual_totals(self, records: List[PayrollRecordDB]) -> Dict:
        """Calculate annual totals from payroll records"""
        
        totals = {
            'total_gross': Decimal('0'),
            'total_hours': Decimal('0'),
            'regular_hours': Decimal('0'),
            'overtime_hours': Decimal('0'),
            'swedish_tax': Decimal('0'),
            'finnish_tax': Decimal('0'),
            'pension': Decimal('0'),
            'health': Decimal('0'),
            'tax_free': Decimal('0'),
            'total_net': Decimal('0')
        }
        
        for record in records:
            totals['total_gross'] += record.gross_salary
            totals['swedish_tax'] += Decimal(str(record.swedish_tax))
            totals['finnish_tax'] += abs(record.finnish_tax_withholding)
            totals['pension'] += abs(record.pension_insurance)
            totals['health'] += abs(record.health_insurance)
            totals['tax_free'] += record.tax_free_portion
            totals['total_net'] += record.net_payment
            
            # Calculate hours from salary items
            for item in record.salary_items:
                if item.quantity:
                    if item.code == '12101':
                        totals['regular_hours'] += Decimal(str(item.quantity))
                    elif item.code in ['12101_2', '12102']:
                        totals['overtime_hours'] += Decimal(str(item.quantity))
        
        totals['total_hours'] = totals['regular_hours'] + totals['overtime_hours']
        
        return totals