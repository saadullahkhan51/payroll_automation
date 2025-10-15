import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
from typing import List
from decimal import Decimal
from database.repository import PayrollRepository
from config.settings import OUTPUT_DIR

class AnnualSummaryGenerator:
    """Generate annual summary for all workers"""
    
    def __init__(self, repository: PayrollRepository):
        self.repo = repository
        self.output_dir = OUTPUT_DIR / "annual"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_all_workers_annual_summary(self, year: int) -> str:
        """Generate annual summary for all workers"""
        
        # Get all employees
        employees = self.repo.get_all_employees()
        
        if not employees:
            raise ValueError(f"No employees found")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year} All Workers"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        
        # Define styles
        bold_font = Font(bold=True)
        header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header row
        row = 1
        ws[f'A{row}'] = "Hlö / person"
        ws[f'B{row}'] = "(1) Rahapalkka kaikki v. 2024 EUR"
        ws[f'C{row}'] = "(2) Rahapalkka Suomi EUR"
        ws[f'D{row}'] = "(3) Rahapalkka Norja EUR"
        ws[f'E{row}'] = "(4) Rahapalkka Ruotsi EUR (sis pvr! ilmoitetaan ruotsin rahapalkkasi)"
        ws[f'F{row}'] = "(5) Ennakonpidätys maksettu ulkomaille EUR (lieto tulorekisteristä)"
        
        # Second header row
        row = 2
        ws[f'B{row}'] = "Amount of all cash wages 2024 EUR"
        ws[f'C{row}'] = "Amount of cash wages from Fin Work 2024 EUR"
        ws[f'D{row}'] = "Amount of cash wages from Norwegian Work 2024 EUR"
        ws[f'E{row}'] = "Amount of cash wages from Sweden Work 2024 EUR"
        ws[f'F{row}'] = "Amount of taxes paid to abroad 2024 EUR"
        
        # Apply header styling
        for r in [1, 2]:
            for col in range(1, 7):
                cell = ws.cell(row=r, column=col)
                cell.font = bold_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Data rows
        row = 3
        
        for employee in employees:
            records = self.repo.get_payroll_records(employee.id, year)
            
            if not records:
                continue
            
            # Calculate totals
            total_gross = sum(r.gross_salary for r in records)
            total_swedish_tax = sum(Decimal(str(r.swedish_tax)) for r in records)
            
            # For this implementation:
            # - All work is considered Swedish work (column E)
            # - Column B (total) = Column E (Swedish)
            # - Columns C (Finland) and D (Norway) = 0
            
            ws[f'A{row}'] = employee.name
            ws[f'B{row}'] = float(total_gross)
            ws[f'C{row}'] = 0  # No Finnish work
            ws[f'D{row}'] = 0  # No Norwegian work
            ws[f'E{row}'] = float(total_gross)  # All Swedish work
            ws[f'F{row}'] = float(total_swedish_tax)
            
            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
            
            # Number formatting
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
            
            row += 1
        
        # Generate filename
        filename = f"annual_summary_all_workers_{year}.xlsx"
        filepath = self.output_dir / filename
        
        # Save workbook
        wb.save(filepath)
        
        return str(filepath)