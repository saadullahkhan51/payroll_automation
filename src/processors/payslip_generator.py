import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from decimal import Decimal
from models.payroll import PayrollRecord
from config.settings import OUTPUT_DIR, TEMPLATE_DIR

class PayslipGenerator:
    """Generate individual payslip Excel files"""
    
    def __init__(self):
        self.output_dir = OUTPUT_DIR / "payslips"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate(self, payroll_record: PayrollRecord) -> str:
        """Generate payslip Excel file"""
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Palkkalaskelma"
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        # Define styles
        header_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header section
        row = 1
        ws.merge_cells(f'C{row}:D{row}')
        ws[f'C{row}'] = "PALKKALASKELMA"
        ws[f'C{row}'].font = header_font
        
        ws[f'D{row + 1}'] = f"Numero: {payroll_record.employee_id}"
        
        row = 2
        ws[f'A{row}'] = f"Lähettäjä: Asiakasyritys Oy, Meritullentie 12, 02110 Espoo"
        ws[f'C{row}'] = "Palkkakausi"
        ws[f'D{row}'] = f"{payroll_record.pay_period_start.strftime('%d.%m.%Y')}-{payroll_record.pay_period_end.strftime('%d.%m.%Y')}"
        
        row = 3
        ws[f'C{row}'] = "Maksupäivä"
        ws[f'D{row}'] = payroll_record.payment_date.strftime('%d/%m/%Y')
        
        row = 4
        ws[f'C{row}'] = "Henkilötunnus:"
        ws[f'D{row}'] = payroll_record.employee_id
        
        row = 5
        ws[f'A{row}'] = f"{payroll_record.name}, {payroll_record.address}"
        ws[f'C{row}'] = "Pankkiyhteys:"
        ws[f'D{row}'] = payroll_record.bank_details
        
        # Tax information section
        row = 12
        ws[f'C{row}'] = "Verotustiedot"
        ws[f'C{row}'].font = bold_font
        
        row = 13
        ws[f'C{row}'] = "Verokortti"
        
        row = 14
        ws[f'C{row}'] = "Perus-%"
        ws[f'D{row}'] = float(payroll_record.tax_info.base_tax_rate)
        
        row = 15
        ws[f'C{row}'] = "Lisä-%"
        ws[f'D{row}'] = float(payroll_record.tax_info.additional_tax_rate)
        
        row = 16
        ws[f'C{row}'] = "Tuloraja/vuosi"
        ws[f'D{row}'] = float(payroll_record.tax_info.income_limit_year)
        
        row = 17
        ws[f'C{row}'] = "Ansio verokauden alusta"
        ws[f'D{row}'] = float(payroll_record.year_to_date_gross)
        
        row = 18
        ws[f'C{row}'] = "Enn.pid. verok. alusta"
        
        # Salary details table
        row = 22
        ws[f'A{row}'] = "Palkkalaji"
        ws[f'B{row}'] = "Selite"
        ws[f'C{row}'] = "Aika/Määrä"
        ws[f'D{row}'] = "A-hinta"
        ws[f'E{row}'] = "Summa"
        
        for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}', f'E{row}']:
            ws[cell].font = bold_font
            ws[cell].border = thin_border
        
        row = 23
        
        # Add salary items
        for item in payroll_record.salary_items:
            ws[f'A{row}'] = item.code
            ws[f'B{row}'] = item.description
            ws[f'C{row}'] = float(item.quantity) if item.quantity else ""
            ws[f'D{row}'] = float(item.rate) if item.rate else ""
            ws[f'E{row}'] = float(item.total)
            row += 1
        
        # Add deductions
        for deduction in payroll_record.deductions:
            ws[f'A{row}'] = deduction.code
            ws[f'B{row}'] = deduction.description
            ws[f'E{row}'] = float(deduction.amount)
            row += 1
        
        # Add benefits
        for benefit in payroll_record.benefits:
            ws[f'A{row}'] = benefit.code
            ws[f'B{row}'] = benefit.description
            ws[f'C{row}'] = float(benefit.quantity) if benefit.quantity else ""
            ws[f'D{row}'] = float(benefit.rate) if benefit.rate else ""
            ws[f'E{row}'] = float(benefit.total)
            row += 1
        
        # Other deductions section
        row += 1
        ws[f'A{row}'] = "Muut palkkatiedot"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        ws[f'A{row}'] = "Ennakonpidätys"
        ws[f'E{row}'] = float(payroll_record.tax_withholding)
        row += 1
        
        ws[f'A{row}'] = "TyEL"
        ws[f'E{row}'] = float(payroll_record.pension_insurance)
        row += 1
        
        ws[f'A{row}'] = "TVM"
        row += 1
        
        ws[f'A{row}'] = f"Sairausvakuutuksen päiväraha (0,84%): {float(payroll_record.health_insurance_daily):.2f}"
        row += 1
        
        ws[f'A{row}'] = "Veroton osuus"
        ws[f'E{row}'] = float(payroll_record.tax_free_portion)
        row += 2
        
        # Net payment
        ws[f'A{row}'] = "Maksetaan"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'E{row}'] = f"{float(payroll_record.net_payment):.2f} €"
        ws[f'E{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Payment certificate
        ws[f'A{row}'] = "PALKKATODISTUS"
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = "Kausi"
        ws[f'B{row}'].font = bold_font
        ws[f'C{row}'] = "Kuluva vuosi"
        ws[f'C{row}'].font = bold_font
        row += 1
        
        ws[f'A{row}'] = "Rahapalkka"
        ws[f'B{row}'] = float(payroll_record.gross_salary)
        ws[f'C{row}'] = float(payroll_record.year_to_date_gross)
        row += 1
        
        ws[f'A{row}'] = "Luontoisetu"
        row += 1
        
        ws[f'A{row}'] = "Enn.al.ansio"
        ws[f'B{row}'] = float(payroll_record.gross_salary)
        ws[f'C{row}'] = float(payroll_record.year_to_date_gross)
        row += 1
        
        ws[f'A{row}'] = "Veroton osuus"
        ws[f'B{row}'] = float(payroll_record.tax_free_portion)
        ws[f'C{row}'] = float(payroll_record.year_to_date_tax_free)
        row += 1
        
        ws[f'A{row}'] = "Enn.pid."
        
        # Generate filename
        filename = f"{payroll_record.employee_id}_{payroll_record.pay_period_start.year}_{payroll_record.pay_period_start.month:02d}_payslip.xlsx"
        filepath = self.output_dir / filename
        
        # Save workbook
        wb.save(filepath)
        
        return str(filepath)
