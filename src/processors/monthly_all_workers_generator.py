import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
from typing import List, Dict
from decimal import Decimal
from datetime import date
from models.payroll import PayrollRecord
from config.settings import OUTPUT_DIR
import calendar

class MonthlyAllWorkersGenerator:
    """Generate monthly salary info for all workers (structured by pay periods)"""
    
    def __init__(self):
        self.output_dir = OUTPUT_DIR / "monthly"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.SEK_TO_EUR = Decimal('0.086')  # Exchange rate
    
    def generate(self, payroll_records: List[PayrollRecord], year: int, month: int) -> str:
        """Generate monthly all workers salary info sheet"""
        
        # Group records by pay period
        # Assume first half (1-15) and second half (16-31) of month
        first_half = []
        second_half = []
        
        for record in payroll_records:
            if record.pay_period_start.day <= 15:
                first_half.append(record)
            else:
                second_half.append(record)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year}-{month:02d} All Workers"
        
        # Define styles
        bold_font = Font(bold=True)
        red_font = Font(color="FF0000", italic=True)
        red_bold_font = Font(color="FF0000", bold=True)
        header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # ====================================================================
        # HEADERS ROW - ALL SECTIONS ON SAME ROW
        # ====================================================================
                
        # Row 2: Red italic subtitles
        subtitle_row = 2
        month_range = lambda m, y: f"1.-{calendar.monthrange(y, m)[1]}."

        ws.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=4)
        ws[f'A{subtitle_row}'] = f"Ruotsin palkat, maksupäivä {month_range(month, year)}{month}.{year}"
        ws[f'A{subtitle_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{subtitle_row}'].font = red_font

        ws.merge_cells(start_row=subtitle_row, start_column=7, end_row=subtitle_row, end_column=16)
        ws[f'G{subtitle_row}'] = f"15. pv"
        ws[f'G{subtitle_row}'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=subtitle_row, start_column=17, end_row=subtitle_row, end_column=26)
        ws[f'Q{subtitle_row}'] = f"{month_range(month, year)[-3:]} pv"
        ws[f'Q{subtitle_row}'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=subtitle_row, start_column=28, end_row=subtitle_row, end_column=36)
        ws[f'AB{subtitle_row}'] = "YHTEENSÄ"
        ws[f'AB{subtitle_row}'].font = bold_font
        ws[f'AB{subtitle_row}'].alignment = Alignment(horizontal='center', vertical='center')

        
        # Row 3: Column headers (all on same row)
        header_row = 3
        
        # First period headers (A-P)
        headers_first = [
            'Nimi', 'Kansalaisuus', 'Suomen\nveronumero', 'Suomen hetu', 'Ruotsin ID',
            '', 'Toitä Ruotsissa', 'Laskelma nro\nSuomi', 'Laskelma nro\nRuotsi',
            'Päivärahat', 'Ansiot EUR', 'Ansiot SEK', 'Päivärahat SEK',
            'Ansiot SEK yht.', 'Vero -%', 'Vero €'
        ]
        
        for col_idx, header in enumerate(headers_first, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Second period headers (Q-AA) - shifted to avoid overlap
        headers_second = [
            'Toitä Ruotsissa', 'Laskelma nro\nSuomi', 'Laskelma nro\nRuotsi',
            'Päivärahat', 'Ansiot EUR', 'Ansiot SEK', 'Päivärahat SEK',
            'Ansiot SEK yht.', 'Vero -%', 'Vero €',''
        ]
        
        for col_idx, header in enumerate(headers_second, start=17):  # Start at Q (column 17)
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # YHTEENSÄ section headers (AB-AH) - shifted to avoid overlap
        yhteensa_headers = [
            ('AB', 'km-korvauksia'),
            ('AC', 'Ansiot SEK'),
            ('AD', 'Päivärahat SEK'),
            ('AE', 'Ansiot SEK yht.'),
            ('AF', 'Vero -%'),
            ('AG', ''),
            ('AH', 'Vero SEK'),
            ('AI', 'Vero €'),
            ('AJ', 'palkkatietoilmoitus\nlomake')
        ]
        
        for col, header in yhteensa_headers:
            cell = ws[f'{col}{header_row}']
            cell.value = header
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        # ====================================================================
        # DATA ROWS
        # ====================================================================
        row = 4
        
        # Write data for both periods on same rows
        if first_half or second_half:
            row = self._write_combined_period_data(
                ws, row, first_half, second_half, 
                thin_border, yellow_fill, red_bold_font
            )
            
            row += 2  # Spacing
        
        # ====================================================================
        # YHTEENSÄ (COMBINED TOTALS) SECTION
        # ====================================================================
        self._write_combined_totals_inline(ws, first_half + second_half, thin_border, yellow_fill, bold_font)
        
        # Move to bottom notes section
        row += 3
        ws[f'A{row}'] = "Removal of Finnish income taxes"
        row += 1
        ws[f'A{row}'] = "Swedish taxes 30% from taxable income"
        row += 1
        ws[f'A{row}'] = "Swedish taxes 30% from allowance"
        row += 2
        ws[f'A{row}'] = "Ruotsin verot 30% verotettavasta tulosta SEK"
        ws[f'A{row}'].font = bold_font
        row += 1
        ws[f'A{row}'] = "Ruotsin verot 30% verotettavista kulukorvauksista SEK"
        ws[f'A{row}'].font = Font(italic=True)
        row += 1
        ws[f'A{row}'] = "Suomen verokannalla lasketun ennakonpidätyksen poisto"
        
        # Set column widths
        col_widths = {
            'A': 15, 'B': 15, 'C': 15, 'D': 13, 'E': 12,
            'F': 5, 'G': 12, 'H': 12, 'I': 12, 'J': 12,
            'K': 12, 'L': 12, 'M': 13, 'N': 13, 'O': 10,
            'P': 12, 'Q': 12, 'R': 12, 'S': 12, 'T': 12,
            'U': 12, 'V': 13, 'W': 13, 'X': 10, 'Y': 12,
            'Z': 12, 'AA': 12, 'AB': 13, 'AC': 13, 'AD': 10,
            'AE': 12, 'AF': 12, 'AG': 12
        }
        
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        
        # Generate filename
        filename = f"monthly_all_workers_{year}_{month:02d}.xlsx"
        filepath = self.output_dir / filename
        
        # Save workbook
        wb.save(filepath)
        
        return str(filepath)
    
    def _write_combined_period_data(self, ws, start_row, first_half, second_half, thin_border, yellow_fill, red_bold_font):
        """Write data for both periods on the same rows"""
        row = start_row
        
        # Determine max number of employees
        max_employees = max(len(first_half) if first_half else 0, len(second_half) if second_half else 0)
        
        # Period totals
        first_totals = {
            'ansiot_eur': Decimal('0'),
            'ansiot_sek': Decimal('0'),
            'paivarahat_sek': Decimal('0'),
            'ansiot_sek_yht': Decimal('0'),
            'vero_eur': Decimal('0')
        }
        
        second_totals = {
            'ansiot_eur': Decimal('0'),
            'ansiot_sek': Decimal('0'),
            'paivarahat_sek': Decimal('0'),
            'ansiot_sek_yht': Decimal('0'),
            'vero_eur': Decimal('0')
        }
        
        # Write data rows
        for i in range(max_employees):
            # First period data (columns A-P)
            if first_half and i < len(first_half):
                record = first_half[i]
                self._write_single_record(ws, row, record, thin_border, yellow_fill, red_bold_font, 1, first_totals)
            
            # Second period data (columns Q-AA)
            if second_half and i < len(second_half):
                record = second_half[i]
                self._write_single_record(ws, row, record, thin_border, yellow_fill, red_bold_font, 17, second_totals)
            
            row += 1
        
        # Write period totals if multiple records
        if (first_half and len(first_half) > 1) or (second_half and len(second_half) > 1):
            if first_half and len(first_half) > 1:
                # First period totals
                ws[f'K{row}'] = float(first_totals['ansiot_eur'])
                ws[f'L{row}'] = float(first_totals['ansiot_sek'])
                ws[f'M{row}'] = float(first_totals['paivarahat_sek'])
                ws[f'N{row}'] = float(first_totals['ansiot_sek_yht'])
                ws[f'P{row}'] = float(first_totals['vero_eur'])
                
                for col in ['K', 'L', 'M', 'N', 'P']:
                    ws[f'{col}{row}'].number_format = '#,##0.00'
                    ws[f'{col}{row}'].border = thin_border
                    ws[f'{col}{row}'].font = Font(bold=True)
            
            if second_half and len(second_half) > 1:
                # Second period totals (columns Q-AA)
                ws[f'U{row}'] = float(second_totals['ansiot_eur'])
                ws[f'V{row}'] = float(second_totals['ansiot_sek'])
                ws[f'W{row}'] = float(second_totals['paivarahat_sek'])
                ws[f'X{row}'] = float(second_totals['ansiot_sek_yht'])
                ws[f'Z{row}'] = float(second_totals['vero_eur'])
                
                for col in ['U', 'V', 'W', 'X', 'Z']:
                    ws[f'{col}{row}'].number_format = '#,##0.00'
                    ws[f'{col}{row}'].border = thin_border
                    ws[f'{col}{row}'].font = Font(bold=True)
            
            row += 1
        
        return row
    
    def _write_single_record(self, ws, row, record, thin_border, yellow_fill, red_bold_font, start_col, totals):
        """Write a single record's data"""
        
        # Calculate values
        ansiot_eur = record.gross_salary - record.tax_free_portion
        ansiot_sek = ansiot_eur / self.SEK_TO_EUR
        paivarahat_sek = record.tax_free_portion / self.SEK_TO_EUR
        ansiot_sek_yht = ansiot_sek + paivarahat_sek
        swedish_tax_eur = sum(
            abs(d.amount) for d in record.deductions 
            if 'Swedish' in d.description or 'Ruotsin' in d.description
        )
        
        # Calculation numbers
        calc_num_ruotsi = 136445 + (record.pay_period_start.month * 100)
        
        # Number of per diem days
        paivarahat_count = 5 if record.tax_free_portion > 0 else 0
        
        # Extract name parts
        name_parts = record.name.split()
        short_name = f"{name_parts[0]} {name_parts[-1]}" if len(name_parts) > 1 else record.name
        
        if start_col == 1:  # First period (columns A-P)
            ws[f'A{row}'] = short_name
            ws[f'B{row}'] = "FI"
            ws[f'C{row}'] = ""  # Finnish tax number
            ws[f'D{row}'] = ""  # Finnish personal ID
            ws[f'E{row}'] = ""  # Swedish ID
            ws[f'F{row}'] = ""
            ws[f'G{row}'] = "x"
            ws[f'H{row}'] = ""  # Calc num Finland
            ws[f'I{row}'] = calc_num_ruotsi
            ws[f'J{row}'] = paivarahat_count if paivarahat_count > 0 else ""
            ws[f'K{row}'] = float(ansiot_eur)
            ws[f'L{row}'] = float(ansiot_sek)
            ws[f'M{row}'] = float(paivarahat_sek)
            ws[f'N{row}'] = float(ansiot_sek_yht)
            ws[f'O{row}'] = "30%"
            ws[f'O{row}'].font = red_bold_font
            ws[f'P{row}'] = float(swedish_tax_eur)
            ws[f'P{row}'].fill = yellow_fill
            
            # Borders
            for col in range(1, 17):
                ws.cell(row=row, column=col).border = thin_border
            
            # Number formatting
            for col in ['K', 'L', 'M', 'N', 'P']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
            
        else:  # Second period (columns Q-AA)
            ws[f'Q{row}'] = "x"  # Toitä Ruotsissa
            ws[f'R{row}'] = ""  # Calc num Finland
            ws[f'S{row}'] = calc_num_ruotsi
            ws[f'T{row}'] = paivarahat_count if paivarahat_count > 0 else ""
            ws[f'U{row}'] = float(ansiot_eur)
            ws[f'V{row}'] = float(ansiot_sek)
            ws[f'W{row}'] = float(paivarahat_sek)
            ws[f'X{row}'] = float(ansiot_sek_yht)
            ws[f'Y{row}'] = "30%"
            ws[f'Y{row}'].font = red_bold_font
            ws[f'Z{row}'] = float(swedish_tax_eur)
            ws[f'Z{row}'].fill = yellow_fill
            
            # Borders
            for col in range(17, 27):
                ws.cell(row=row, column=col).border = thin_border
            
            # Number formatting
            for col in ['U', 'V', 'W', 'X', 'Z']:
                ws[f'{col}{row}'].number_format = '#,##0.00'
        
        # Update totals
        totals['ansiot_eur'] += ansiot_eur
        totals['ansiot_sek'] += ansiot_sek
        totals['paivarahat_sek'] += paivarahat_sek
        totals['ansiot_sek_yht'] += ansiot_sek_yht
        totals['vero_eur'] += Decimal(str(swedish_tax_eur))
    
    def _write_combined_totals_inline(self, ws, all_records, thin_border, yellow_fill, bold_font):
        """Write YHTEENSÄ combined totals in columns AB-AJ for each worker"""
        
        # Group records by employee
        employee_records = {}
        for record in all_records:
            if record.employee_id not in employee_records:
                employee_records[record.employee_id] = []
            employee_records[record.employee_id].append(record)
        
        # Write individual employee totals starting at row 4
        row = 4
        
        # Grand totals across all employees
        grand_totals = {
            'km_korvauksia': Decimal('0'),
            'ansiot_sek': Decimal('0'),
            'paivarahat_sek': Decimal('0'),
            'ansiot_sek_yht': Decimal('0'),
            'vero_sek': Decimal('0'),
            'vero_eur': Decimal('0')
        }
        
        # Calculate and write totals for each employee
        for employee_id, records in employee_records.items():
            employee_totals = {
                'km_korvauksia': Decimal('0'),
                'ansiot_sek': Decimal('0'),
                'paivarahat_sek': Decimal('0'),
                'ansiot_sek_yht': Decimal('0'),
                'vero_sek': Decimal('0'),
                'vero_eur': Decimal('0')
            }
            
            # Sum up both periods for this employee
            for record in records:
                # Calculate earnings (gross minus tax-free)
                ansiot_eur = record.gross_salary - record.tax_free_portion
                ansiot_sek = ansiot_eur / self.SEK_TO_EUR
                
                # Per diem in SEK
                paivarahat_sek = record.tax_free_portion / self.SEK_TO_EUR
                
                # Travel compensation (from benefits)
                travel_comp = sum(
                    b.total for b in record.benefits 
                    if 'MK' in b.code or 'matka' in b.description.lower()
                )
                
                # Swedish tax
                swedish_tax_eur = sum(
                    abs(d.amount) for d in record.deductions 
                    if 'Swedish' in d.description or 'Ruotsin' in d.description
                )
                swedish_tax_sek = Decimal(str(swedish_tax_eur)) / self.SEK_TO_EUR
                
                # Accumulate employee totals
                employee_totals['km_korvauksia'] += travel_comp
                employee_totals['ansiot_sek'] += ansiot_sek
                employee_totals['paivarahat_sek'] += paivarahat_sek
                employee_totals['ansiot_sek_yht'] += ansiot_sek + paivarahat_sek
                employee_totals['vero_sek'] += swedish_tax_sek
                employee_totals['vero_eur'] += Decimal(str(swedish_tax_eur))
            
            # Write this employee's totals
            # AB: km-korvauksia (travel compensation)
            if employee_totals['km_korvauksia'] > 0:
                ws[f'AB{row}'] = float(employee_totals['km_korvauksia'])
                ws[f'AB{row}'].number_format = '#,##0.00'
                ws[f'AB{row}'].border = thin_border
            
            # AC: Ansiot SEK
            ws[f'AC{row}'] = float(employee_totals['ansiot_sek'])
            ws[f'AC{row}'].number_format = '#,##0.00'
            ws[f'AC{row}'].border = thin_border
            
            # AD: Päivärahat SEK
            ws[f'AD{row}'] = float(employee_totals['paivarahat_sek'])
            ws[f'AD{row}'].number_format = '#,##0.00'
            ws[f'AD{row}'].border = thin_border
            
            # AE: Ansiot SEK yht. (combined total)
            ws[f'AE{row}'] = float(employee_totals['ansiot_sek_yht'])
            ws[f'AE{row}'].number_format = '#,##0.00'
            ws[f'AE{row}'].border = thin_border
            
            # AF: Vero -% (RED 30%)
            ws[f'AF{row}'] = "30%"
            ws[f'AF{row}'].font = Font(color="FF0000", bold=True)
            ws[f'AF{row}'].border = thin_border
            ws[f'AF{row}'].alignment = Alignment(horizontal='center')
            
            # AG: Empty column
            ws[f'AG{row}'] = ""
            ws[f'AG{row}'].border = thin_border
            
            # AH: Vero SEK
            ws[f'AH{row}'] = float(employee_totals['vero_sek'])
            ws[f'AH{row}'].number_format = '#,##0.00'
            ws[f'AH{row}'].border = thin_border
            
            # AI: Vero €
            ws[f'AI{row}'] = float(employee_totals['vero_eur'])
            ws[f'AI{row}'].number_format = '#,##0.00'
            ws[f'AI{row}'].border = thin_border
            
            # AJ: palkkatietoilmoitus lomake
            ws[f'AJ{row}'] = "x"
            ws[f'AJ{row}'].border = thin_border
            ws[f'AJ{row}'].alignment = Alignment(horizontal='center')
            
            # Accumulate to grand totals
            for key in grand_totals:
                grand_totals[key] += employee_totals[key]
            
            row += 1
        
        # Write grand total row (sum of all employees)
        # AE: Grand total (YELLOW)
        ws[f'AE{row}'] = float(grand_totals['ansiot_sek_yht'])
        ws[f'AE{row}'].number_format = '#,##0.00'
        ws[f'AE{row}'].fill = yellow_fill
        ws[f'AE{row}'].border = thin_border
        ws[f'AE{row}'].font = bold_font
        
        # AF: 30%
        ws[f'AF{row}'] = "30%"
        ws[f'AF{row}'].font = Font(color="FF0000", bold=True)
        ws[f'AF{row}'].border = thin_border
        
        # AH: Vero SEK grand total (YELLOW)
        ws[f'AH{row}'] = float(grand_totals['vero_sek'])
        ws[f'AH{row}'].number_format = '#,##0.00'
        ws[f'AH{row}'].fill = yellow_fill
        ws[f'AH{row}'].border = thin_border
        ws[f'AH{row}'].font = bold_font
        
        # AI: Vero € grand total
        ws[f'AI{row}'] = float(grand_totals['vero_eur'])
        ws[f'AI{row}'].number_format = '#,##0.00'
        ws[f'AI{row}'].border = thin_border
        ws[f'AI{row}'].font = bold_font
