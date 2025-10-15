import sys
from pathlib import Path
from datetime import datetime, date

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / 'src'))

from api.mock_likeit import MockLikeitAPI
from processors.monthly_all_workers_generator import MonthlyAllWorkersGenerator
from database.db import init_db, SessionLocal
from database.repository import PayrollRepository
from models.employee import Employee
from models.payroll import PayrollRecord, TaxInfo, SalaryItem, Deduction
from decimal import Decimal

def create_test_records_for_month(api: MockLikeitAPI, year: int, month: int):
    """Create test records with two pay periods for the month"""
    
    employees = api.get_all_employees()
    all_records = []
    
    # Create records for first period (1-15)
    for emp in employees:
        record = api.get_employee_payroll(emp['employee_id'], year, month)
        # Adjust to first period
        record.pay_period_start = date(year, month, 1)
        record.pay_period_end = date(year, month, 15)
        record.payment_date = date(year, month, 20)
        all_records.append(record)
    
    # Create records for second period (16-31)
    for emp in employees:
        record = api.get_employee_payroll(emp['employee_id'], year, month)
        # Adjust to second period
        if month == 12:
            last_day = 31
        elif month in [4, 6, 9, 11]:
            last_day = 30
        elif month == 2:
            last_day = 28 if year % 4 != 0 else 29
        else:
            last_day = 31
        
        record.pay_period_start = date(year, month, 16)
        record.pay_period_end = date(year, month, last_day)
        record.payment_date = date(year, month, last_day)
        all_records.append(record)
    
    return all_records

def print_section(title):
    """Print a formatted section header"""
    print("\n" + "=" * 80)
    print(f"  {title}")
    print("=" * 80)

def print_success(message):
    """Print success message"""
    print(f"  ✓ {message}")

def print_info(message):
    """Print info message"""
    print(f"  → {message}")

def print_error(message):
    """Print error message"""
    print(f"  ✗ {message}")

def test_monthly_all_workers_generator():
    """Test the monthly all workers generator"""
    
    start_time = datetime.now()
    
    print("\n" + "╔" + "═" * 78 + "╗")
    print("║" + " " * 15 + "MONTHLY ALL WORKERS GENERATOR - TEST" + " " * 22 + "║")
    print("╚" + "═" * 78 + "╝")
    
    # Configuration
    year = 2025
    test_month = 8  # August
    
    try:
        # ====================================================================
        # STEP 1: Initialize
        # ====================================================================
        print_section("STEP 1: Initialization")
        
        # Initialize database
        init_db()
        db = SessionLocal()
        repo = PayrollRepository(db)
        print_success("Database initialized")
        
        # Initialize API
        api = MockLikeitAPI()
        employees = api.get_all_employees()
        print_success(f"Mock API initialized with {len(employees)} employees")
        
        for emp in employees:
            print_info(f"Employee: {emp['name']} ({emp['employee_id']})")
        
        # ====================================================================
        # STEP 2: Generate Test Data
        # ====================================================================
        print_section(f"STEP 2: Generating Payroll Data for {test_month}/{year}")
        
        # Create records with two pay periods
        print_info("Creating records for first period (1-15)...")
        print_info("Creating records for second period (16-31)...")
        
        all_records = create_test_records_for_month(api, year, test_month)
        
        print_success(f"Generated {len(all_records)} payroll records")
        print_info(f"  - First period (1-15): {len(all_records)//2} records")
        print_info(f"  - Second period (16-31): {len(all_records)//2} records")
        
        # Save to database
        print("\n  Saving to database...")
        for record in all_records:
            # Save employee
            employee = Employee(
                employee_id=record.employee_id,
                name=record.name,
                address=record.address,
                bank_details=record.bank_details
            )
            repo.save_employee(employee)
            
            # Save payroll record
            repo.save_payroll_record(record)
        
        print_success("All records saved to database")
        
        # ====================================================================
        # STEP 3: Display Sample Data
        # ====================================================================
        print_section("STEP 3: Sample Record Details")
        
        sample = all_records[0]
        print_info(f"Employee: {sample.name}")
        print_info(f"Period: {sample.pay_period_start} to {sample.pay_period_end}")
        print_info(f"Gross Salary: {sample.gross_salary:.2f} €")
        print_info(f"Tax-Free Portion: {sample.tax_free_portion:.2f} €")
        print_info(f"Net Payment: {sample.net_payment:.2f} €")
        
        # Calculate Swedish tax
        swedish_tax = sum(
            abs(d.amount) for d in sample.deductions 
            if 'Swedish' in d.description or 'Ruotsin' in d.description
        )
        print_info(f"Swedish Tax: {swedish_tax:.2f} €")
        
        # ====================================================================
        # STEP 4: Generate Monthly All Workers Report
        # ====================================================================
        print_section(f"STEP 4: Generating Monthly All Workers Report")
        
        generator = MonthlyAllWorkersGenerator()
        
        print_info("Generating report...")
        print_info("  - Processing first period (1-15)...")
        print_info("  - Processing second period (16-31)...")
        print_info("  - Calculating combined totals (YHTEENSÄ)...")
        
        filepath = generator.generate(all_records, year, test_month)
        
        print_success(f"Report generated successfully!")
        print_info(f"File: {Path(filepath).name}")
        print_info(f"Location: {filepath}")
        
        # ====================================================================
        # STEP 5: Verify File Contents
        # ====================================================================
        print_section("STEP 5: File Verification")
        
        if Path(filepath).exists():
            file_size = Path(filepath).stat().st_size / 1024  # KB
            print_success(f"File exists: {file_size:.2f} KB")
            
            # Try to open and verify structure
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                
                print_success(f"Workbook opened successfully")
                print_info(f"Sheet name: {ws.title}")
                print_info(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
                
                # Check key cells
                print("\n  Checking key cells:")
                
                # Period 1 code
                if ws['A1'].value:
                    print_info(f"A1 (Period 1 code): {ws['A1'].value}")
                
                # Period 1 subtitle
                if ws['B2'].value:
                    print_info(f"B2 (Subtitle): {ws['B2'].value[:50]}...")
                
                # First employee name
                first_name_row = 4  # Assuming headers at row 3
                if ws[f'A{first_name_row}'].value:
                    print_info(f"A{first_name_row} (First employee): {ws[f'A{first_name_row}'].value}")
                
                # Check YHTEENSÄ section
                yhteensa_found = False
                for row in range(1, ws.max_row + 1):
                    if ws[f'AC{row}'].value == "YHTEENSÄ":
                        print_info(f"AC{row} (YHTEENSÄ found at row {row})")
                        yhteensa_found = True
                        break
                
                if yhteensa_found:
                    print_success("YHTEENSÄ section found")
                else:
                    print_error("YHTEENSÄ section not found")
                
                wb.close()
                
            except Exception as e:
                print_error(f"Error verifying file: {str(e)}")
        else:
            print_error("File does not exist!")
        
        # ====================================================================
        # STEP 6: Summary Statistics
        # ====================================================================
        print_section("STEP 6: Summary Statistics")
        
        # Calculate totals
        total_gross = sum(r.gross_salary for r in all_records)
        total_tax_free = sum(r.tax_free_portion for r in all_records)
        total_net = sum(r.net_payment for r in all_records)
        
        total_swedish_tax = sum(
            sum(abs(d.amount) for d in r.deductions 
                if 'Swedish' in d.description or 'Ruotsin' in d.description)
            for r in all_records
        )
        
        print_info(f"Total Employees: {len(employees)}")
        print_info(f"Total Pay Periods: 2 (1-15, 16-31)")
        print_info(f"Total Records: {len(all_records)}")
        print_info(f"Total Gross Salary: {total_gross:.2f} €")
        print_info(f"Total Tax-Free Portion: {total_tax_free:.2f} €")
        print_info(f"Total Swedish Tax: {total_swedish_tax:.2f} €")
        print_info(f"Total Net Payment: {total_net:.2f} €")
        
        # SEK conversions
        SEK_TO_EUR = Decimal('0.086')
        total_ansiot_sek = (total_gross - total_tax_free) / SEK_TO_EUR
        total_paivarahat_sek = total_tax_free / SEK_TO_EUR
        
        print_info(f"\nSEK Conversions:")
        print_info(f"Total Ansiot SEK: {float(total_ansiot_sek):,.2f}")
        print_info(f"Total Päivärahat SEK: {float(total_paivarahat_sek):,.2f}")
        print_info(f"Total Combined SEK: {float(total_ansiot_sek + total_paivarahat_sek):,.2f}")
        
        # ====================================================================
        # Database Cleanup
        # ====================================================================
        db.close()
        
    except Exception as e:
        print_error(f"Critical error: {str(e)}")
        import traceback
        traceback.print_exc()
    
    # ====================================================================
    # FINAL REPORT
    # ====================================================================
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    
    print("\n" + "╔" + "═" * 78 + "╗")
    print("║" + " " * 27 + "TEST COMPLETED!" + " " * 32 + "║")
    print("╚" + "═" * 78 + "╝")
    
    print(f"\n  ⏱️  Execution Time: {duration:.2f} seconds")
        
    print(f"\n  📋 What to Check in the Excel File:")
    print(f"     1. First period section (1-15) with yellow period code")
    print(f"     2. Red italic subtitle 'Ruotsin palkat, maksupäivä...'")
    print(f"     3. Blue header row with all column names")
    print(f"     4. Employee data rows with 'x' in Toitä Ruotsissa column")
    print(f"     5. Calculation numbers (Laskelma nro Ruotsi)")
    print(f"     6. Per diem days count (Päivärahat)")
    print(f"     7. Earnings in EUR and SEK")
    print(f"     8. Red '30%' tax percentage")
    print(f"     9. Yellow highlighted tax amounts in €")
    print(f"     10. Second period section (16-31) starting at column P")
    print(f"     11. YHTEENSÄ section with combined totals at columns Z-AH")
    print(f"     12. Yellow highlights on combined totals")
    print(f"     13. Bottom notes about tax removal")
    
    print(f"\n  🎯 Next Steps:")
    print(f"     1. Open the Excel file and verify the format")
    print(f"     2. Check that it matches the reference images")
    print(f"     3. Verify all calculations are correct")
    print(f"     4. Test with different months if needed")
    
    print("\n" + "=" * 80 + "\n")

if __name__ == "__main__":
    test_monthly_all_workers_generator()